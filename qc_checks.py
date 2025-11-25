import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

DATE_FORMAT = "%Y-%m-%d"

# Excel color styles
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")


# ----------------------------- 1️⃣ Detect Monitoring Period -----------------------------
def detect_period_from_rosco(rosco_path):
    df = pd.read_excel(rosco_path, header=None)
    value_col = df.iloc[:, 1].astype(str)
    period_row = value_col[value_col.str.contains("Monitoring Periods", na=False)]
    if period_row.empty:
        raise ValueError("Could not find 'Monitoring Periods' in Rosco file.")
    text = period_row.iloc[0]
    found = re.findall(r"\d{4}-\d{2}-\d{2}", text)
    if len(found) >= 2:
        start_date = pd.to_datetime(found[0], format=DATE_FORMAT)
        end_date = pd.to_datetime(found[1], format=DATE_FORMAT)
        return start_date, end_date
    else:
        raise ValueError("Could not parse monitoring period dates from Rosco file.")


# ----------------------------- 2️⃣ Load BSR -----------------------------
def detect_header_row(bsr_path):
    df_sample = pd.read_excel(bsr_path, header=None, nrows=200)
    for i, row in df_sample.iterrows():
        row_str = " ".join(row.dropna().astype(str).tolist()).lower()
        if "region" in row_str and "market" in row_str and "broadcaster" in row_str:
            return i
        if "date" in row_str and ("utc" in row_str or "gmt" in row_str):
            return i
    raise ValueError("Could not detect header row in BSR file.")


def load_bsr(bsr_path):
    header_row = detect_header_row(bsr_path)
    df = pd.read_excel(bsr_path, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    return df


# ----------------------------- 3️⃣ Period Check -----------------------------
def period_check(df, start_date, end_date):
    date_col = next((c for c in df.columns if "date" in str(c).lower()), None)
    if not date_col:
        df["Within_Period_OK"] = True
        df["Within_Period_Remark"] = ""
        return df
    df["Date_checked"] = pd.to_datetime(df[date_col], errors="coerce").dt.date
    df["Within_Period_OK"] = df["Date_checked"].between(start_date.date(), end_date.date())
    df["Within_Period_Remark"] = df["Within_Period_OK"].apply(lambda x: "" if x else "Date outside monitoring period")
    return df


# ----------------------------- 4️⃣ Completeness Check -----------------------------
def completeness_check(df):
    keywords = ["channel", "aud", "price", "match"]
    matched_cols = [col for col in df.columns if any(kw in str(col).lower() for kw in keywords)]
    if not matched_cols:
        df["Completeness_OK"] = True
        df["Completeness_Remark"] = ""
        return df
    df["Completeness_OK"] = df[matched_cols].notna().all(axis=1)
    df["Completeness_Remark"] = df["Completeness_OK"].apply(lambda x: "" if x else "Missing key fields")
    return df


# ----------------------------- 5️⃣ Overlap / Duplicate / Day Break -----------------------------
def overlap_duplicate_daybreak_check(df):
    df_result = df.copy()
    channel_col = next((c for c in df.columns if "channel" in str(c).lower()), None)
    start_col = next((c for c in df.columns if "start" in str(c).lower()), None)
    end_col = next((c for c in df.columns if "end" in str(c).lower()), None)
    date_col = next((c for c in df.columns if "date" in str(c).lower()), None)

    for col in [start_col, end_col]:
        if col:
            df_result[col] = pd.to_datetime(df_result[col], errors="coerce")

    overlap_flags = [False] * len(df_result)
    if channel_col and start_col and end_col and date_col:
        df_sorted = df_result.sort_values(by=[channel_col, date_col, start_col]).reset_index(drop=True)
        prev_end = prev_channel = prev_date = None
        for i, row in df_sorted.iterrows():
            overlap = False
            if prev_channel == row[channel_col] and prev_date == row[date_col]:
                if pd.notna(row[start_col]) and pd.notna(prev_end) and row[start_col] < prev_end:
                    overlap = True
            overlap_flags[i] = overlap
            prev_end = row[end_col]
            prev_channel = row[channel_col]
            prev_date = row[date_col]
        df_result["No_Overlap"] = ~pd.Series(overlap_flags, index=df_sorted.index)
    else:
        df_result["No_Overlap"] = True

    df_result["No_Overlap_Remark"] = df_result["No_Overlap"].apply(lambda x: "" if x else "Overlap detected")

    # Duplicate check
    exclude_keywords = ["_ok", "within", "date_checked"]
    dup_cols = [c for c in df_result.columns if not any(x in str(c).lower() for x in exclude_keywords)]
    if dup_cols:
        hashes = pd.util.hash_pandas_object(df_result[dup_cols], index=False)
        df_result["Is_Duplicate"] = hashes.duplicated(keep=False)
    else:
        df_result["Is_Duplicate"] = False
    df_result["Is_Duplicate_OK"] = ~df_result["Is_Duplicate"]
    df_result["Is_Duplicate_Remark"] = df_result["Is_Duplicate"].apply(lambda x: "" if not x else "Duplicate row found")

    # Day break check
    if start_col and end_col:
        df_result["Day_Break_OK"] = ~((df_result[start_col].dt.day != df_result[end_col].dt.day) &
                                      (df_result[start_col].dt.hour >= 20))
    else:
        df_result["Day_Break_OK"] = True
    df_result["Day_Break_Remark"] = df_result["Day_Break_OK"].apply(lambda x: "" if x else "Day break mismatch")
    return df_result


# ----------------------------- 6️⃣ Program Category Check -----------------------------
def program_category_check(df):
    prog_col = next((c for c in df.columns if "type" in str(c).lower() and "program" in str(c).lower()), None)
    dur_col = next((c for c in df.columns if "duration" in str(c).lower()), None)
    if prog_col is None or dur_col is None:
        df["Program_Category_OK"] = True
        df["Program_Category_Remark"] = ""
        return df

    def parse_duration(val):
        if pd.isna(val):
            return None
        val_str = str(val).strip()
        try:
            if ":" in val_str:
                t = pd.to_datetime(val_str, errors="coerce").time()
                return t.hour * 60 + t.minute
            else:
                return float(val_str)
        except Exception:
            return None

    def expected_category(duration_min):
        if duration_min is None:
            return "unknown"
        if duration_min >= 120:
            return "live"
        elif 60 <= duration_min < 120:
            return "repeat"
        elif 30 <= duration_min < 60:
            return "highlights"
        elif 0 < duration_min < 30:
            return "support"
        else:
            return "unknown"

    results, remarks = [], []
    for _, row in df.iterrows():
        prog_val = str(row[prog_col]).strip().lower()
        dur_min = parse_duration(row[dur_col])
        expected = expected_category(dur_min)
        ok = expected in prog_val or prog_val in expected
        results.append(ok)
        remarks.append("" if ok else f"Program type '{prog_val}' does not match duration category '{expected}'")

    df["Program_Category_OK"] = results
    df["Program_Category_Remark"] = remarks
    return df


# ----------------------------- 7️⃣ Duration Check -----------------------------
def duration_check(df):
    """Validate program type vs actual duration (Start (UTC) / End (UTC))."""
    print("\n--- DEBUG: Running Duration Check ---")

    # --- Clean column names ---
    df.columns = [str(c).strip() for c in df.columns]

    # --- Detect columns robustly ---
    start_col = None
    end_col = None
    type_col = None
    for col in df.columns:
        col_l = col.lower().strip()
        if col_l in ["start (utc)", "start"]:
            start_col = col
        elif col_l in ["end (utc)", "end"]:
            end_col = col
        elif "type" in col_l and "program" in col_l:
            type_col = col

    if start_col is None or end_col is None or type_col is None:
        print(f"⚠️  Missing columns. Found Start={start_col}, End={end_col}, Type={type_col}")
        df["Duration_Check_OK"] = True
        df["Expected_Category_From_Duration"] = "unknown"
        return df

    # --- Convert to string to avoid NaT issues ---
    df[start_col] = df[start_col].astype(str).str.strip()
    df[end_col] = df[end_col].astype(str).str.strip()

    # --- Helper: parse HH:MM:SS to minutes ---
    def parse_hms_to_minutes(val):
        if not val or val in ["None", "nan", "NaT"]:
            return None
        try:
            parts = val.split(":")
            if len(parts) >= 2:
                h, m = int(parts[0]), int(parts[1])
                s = int(parts[2]) if len(parts) == 3 else 0
                return h * 60 + m + s / 60
        except Exception as e:
            print(f"[WARN] Could not parse time '{val}': {e}")
        return None

    # --- Helper: classify by duration ---
    def expected_category(duration_min):
        if duration_min is None:
            return "unknown"
        if duration_min >= 120:
            return "live"
        elif 60 <= duration_min < 120:
            return "repeat"
        elif 30 <= duration_min < 60:
            return "highlights"
        elif 0 < duration_min < 30:
            return "support"
        else:
            return "unknown"

    expected_list = []
    ok_list = []

    for idx, row in df.iterrows():
        start_val = row[start_col]
        end_val = row[end_col]
        actual_prog = str(row[type_col]).strip().lower() if pd.notna(row[type_col]) else "unknown"

        start_min = parse_hms_to_minutes(start_val)
        end_min = parse_hms_to_minutes(end_val)

        if start_min is None or end_min is None:
            duration_min = None
        else:
            duration_min = end_min - start_min
            if duration_min < 0:
                duration_min += 24 * 60  # Handle midnight crossover

        expected = expected_category(duration_min)
        ok = expected in actual_prog or actual_prog in expected

        expected_list.append(expected)
        ok_list.append(ok)

        print(f"[Row {idx}] Start={start_val} | End={end_val} | Duration(min)={duration_min} | "
              f"Expected='{expected}' | Actual='{actual_prog}' | OK={ok}")

    df["Expected_Category_From_Duration"] = expected_list
    df["Duration_Check_OK"] = ok_list

    print("--- DEBUG: Duration Check Completed ---\n")
    return df

# 8️⃣ Event / Matchday / Competition Check
def check_event_matchday_competition(df_worksheet, df_data=None, rosco_path=None, debug_rows=20):
    """
    Validate Event / Competition / Matchday / Match combinations.

    Inputs:
      - df_worksheet : DataFrame of the main worksheet (the BSR "Worksheet")
          expected columns: "Competition", "Event", "Matchday", "Home Team", "Away Team", maybe "Match"
      - df_data : optional DataFrame extracted from the 'Data' sheet (the reference/master lists).
      - rosco_path : optional path to Excel; used if df_data is None to try to extract reference values from that file.
      - debug_rows: how many rows to print for debug output

    Output:
      - same df_worksheet with two new columns:
          Event_Matchday_Competition_OK (bool)
          Event_Matchday_Competition_Remark (string)
    """

    # --- Helper: normalize text ---
    def norm(x):
        if pd.isna(x):
            return ""
        return str(x).strip()

    def norm_lower(x):
        return norm(x).lower()

    # --- Get reference competitions / allowed values ---
    reference_comps = set()
    reference_matches = set()  # optional: canonical "home vs away" pairs if available
    reference_matchday_counts = {}  # optional expected counts per (competition, matchday)

    if df_data is None and rosco_path is not None:
        # attempt to load a 'Data' sheet or the first sheet that looks like the data table
        try:
            xls = pd.read_excel(rosco_path, sheet_name=None)
            # try common names
            priority = ["Data", "data", "Monitoring list", "monitoring list", "Monitoring List"]
            found_df = None
            for p in priority:
                if p in xls:
                    found_df = xls[p]
                    break
            if found_df is None:
                # fallback: pick sheet that has words like 'Type of programme' or 'Competition' in header rows
                for name, sheet in xls.items():
                    header_text = " ".join(sheet.columns.astype(str).tolist()).lower()
                    if "competition" in header_text or "type of programme" in header_text or "type of program" in header_text:
                        found_df = sheet
                        break
            if found_df is not None:
                df_data = found_df
        except Exception:
            df_data = None

    # If df_data is available, extract competition names and optional counts
    if isinstance(df_data, pd.DataFrame):
        # strategy: scan df_data content for competition-like strings
        df_tmp = df_data.astype(str).applymap(lambda v: v.strip() if pd.notna(v) else "")
        # collect distinct non-empty strings that look like competition names
        for col in df_tmp.columns:
            for val in df_tmp[col].unique():
                v = str(val).strip()
                if v and v not in ["0", "nan", "-", "None"]:
                    # filter out lines that look numeric counts (only digits)
                    if not re.fullmatch(r"^\d+$", v):
                        reference_comps.add(v.lower())

        # attempt to read counts if present: some Data sheets have count rows above/below the headers
        # Look for numeric entries adjacent to competition names in columns
        # Heuristic: if the first few rows contain digits under the same columns as competition names, store count.
        try:
            # look at the first ~10 rows for numeric counts under columns that are competition names
            for col in df_data.columns:
                numeric_counts = []
                for r in range(min(10, len(df_data))):
                    try:
                        v = df_data.iloc[r][col]
                        if pd.notna(v) and str(v).strip().isdigit():
                            numeric_counts.append(int(str(v).strip()))
                    except Exception:
                        continue
                if numeric_counts:
                    # pick a representative (first) numeric if consistent
                    reference_matchday_counts[col.strip().lower()] = numeric_counts[0]
        except Exception:
            pass

    # fallback: if still empty, use some likely defaults
    if not reference_comps:
        reference_comps = set([
            "bundesliga", "2. bundesliga", "dfb-pokal", "dfl supercup",
            "premier league", "epl", "la liga", "serie a", "champions league"
        ])

    # Precompute a lowercase set for quick lookup
    reference_comps_lower = set(x.lower() for x in reference_comps)

    # --- Prepare output columns ---
    df = df_worksheet.copy()
    df["Event_Matchday_Competition_OK"] = False
    df["Event_Matchday_Competition_Remark"] = ""

    # We'll build grouping counts to verify number of matches per (Competition, Matchday)
    grouped_counts = {}

    # iterate rows
    for idx, row in df.iterrows():
        competition = norm(row.get("Competition", ""))
        event = norm(row.get("Event", ""))
        matchday = norm(row.get("Matchday", ""))

        # some BSRs have 'Matchday' in other column names like 'Matchday ' or 'Match Day' - check alternatives
        if not matchday:
            # try columns similar to matchday
            for c in df.columns:
                if "matchday" in c.lower() or "match day" in c.lower() or c.lower().strip() == "match":
                    matchday = norm(row.get(c, ""))
                    if matchday:
                        break

        # find home/away or match field
        home = norm(row.get("Home Team", "")) or norm(row.get("HomeTeam", "")) or norm(row.get("Home", ""))
        away = norm(row.get("Away Team", "")) or norm(row.get("AwayTeam", "")) or norm(row.get("Away", ""))

        remarks = []
        ok = True

        # 1) Missing fields
        if not competition or competition.strip() in ["-", "nan", "none"]:
            ok = False
            remarks.append("Missing Competition")
        if not event or event.strip() in ["-", "nan", "none"]:
            ok = False
            remarks.append("Missing Event")
        if not matchday or matchday.strip() in ["-", "nan", "none"]:
            ok = False
            remarks.append("Missing Matchday")
        if not (home and away):
            # sometimes matches are in 'Match' or 'Program Title', try match detection
            match_text = norm(row.get("Match", "")) or norm(row.get("Program Title", "")) or norm(row.get("Combined", ""))
            # a simple heuristic: look for ' vs ' or ' v ' separators
            if " vs " in match_text.lower() or " v " in match_text.lower():
                # we accept this as a match, but still prefer to split
                try:
                    parts = re.split(r"\s+v(?:s|)\.?\s+|\s+vs\.?\s+|\s+v\s+", match_text, flags=re.IGNORECASE)
                    if len(parts) >= 2:
                        home = parts[0].strip()
                        away = parts[1].strip()
                except Exception:
                    pass
            else:
                ok = False
                remarks.append("Missing Home/Away or Match field")

        # 2) Validate competition against reference list
        comp_l = competition.lower()
        # some competitions appear with extra words, do a contains check
        comp_matches_reference = False
        for rc in reference_comps_lower:
            if rc and (rc in comp_l or comp_l in rc):
                comp_matches_reference = True
                break
        if not comp_matches_reference:
            ok = False
            remarks.append("Competition not in reference list")

        # 3) Simple event-matchday-match consistency: check if 'matchday' value format looks valid (MD, Round, etc.)
        # Accept common formats: 'Matchday 01', 'MD01', 'Round 01', 'Round 1', 'Matchday 1'
        if matchday:
            if not re.search(r"(matchday|md|round|rd|r|matchday)\s*\d+", matchday.lower()):
                # allow some textual forms like 'Finals', 'Semi', 'Quarter'
                if matchday.lower() not in ["final", "finals", "semi", "semifinal", "quarterfinal", "playoffs", "-"]:
                    # it's not necessarily an error; just add a warning
                    remarks.append("Unusual matchday format")

        # 4) If we have a reference expected counts mapping (from df_data), count per (competition, matchday)
        comp_key = (competition.strip().lower(), matchday.strip().lower())
        grouped_counts.setdefault(comp_key, 0)
        grouped_counts[comp_key] += 1

        # Compose final remark and set OK
        df.at[idx, "Event_Matchday_Competition_OK"] = ok
        df.at[idx, "Event_Matchday_Competition_Remark"] = "; ".join(remarks) if remarks else "OK"

    # 5) If reference_matchday_counts available, compare counts and append remarks for rows belonging to mismatch groups
    # reference_matchday_counts keys may be competition names -> expected counts per matchday (heuristic)
    if reference_matchday_counts:
        # For each group in grouped_counts, compare to reference (best-effort)
        for (comp, mday), observed in grouped_counts.items():
            expected = None
            # try to find matching competition in reference counts map
            for ref_comp_name, cnt in reference_matchday_counts.items():
                if ref_comp_name and (ref_comp_name in comp or comp in ref_comp_name):
                    expected = cnt
                    break
            if expected is not None and observed != expected:
                # flag all rows in df with this (comp, mday)
                mask = df[
                    df.get("Competition", "").astype(str).str.strip().str.lower() == comp
                ]["Competition"].notna()
                # append a remark for each row in this group
                for idx in df[
                    (df.get("Competition", "").astype(str).str.strip().str.lower() == comp) &
                    (df.get("Matchday", "").astype(str).str.strip().str.lower() == mday)
                ].index:
                    prev = df.at[idx, "Event_Matchday_Competition_Remark"]
                    extra = f"Mismatch matches per matchday: expected {expected}, found {observed}"
                    df.at[idx, "Event_Matchday_Competition_Remark"] = (prev + "; " + extra) if prev else extra
                    df.at[idx, "Event_Matchday_Competition_OK"] = False

    # --- Debug prints (first few rows) ---
    print("=== Event/Matchday/Competition QC summary (first rows) ===")
    for idx in range(min(debug_rows, len(df))):
        r = df.iloc[idx]
        print(f"[Row {idx}] Competition='{r.get('Competition','')}' | Event='{r.get('Event','')}' | Matchday='{r.get('Matchday','')}' | "
              f"Home='{r.get('Home Team', r.get('Home', ''))}' Away='{r.get('Away Team', r.get('Away', ''))}' | "
              f"OK={r['Event_Matchday_Competition_OK']} | Remark={r['Event_Matchday_Competition_Remark']}")
    print("=== End summary ===\n")

    return df

# -----------------------------------------------------------
# 9️⃣ Market / Channel / Program / Duration Consistency Check

def market_channel_program_duration_check(df_worksheet, reference_df=None, debug_rows=10):
    df = df_worksheet.copy()
    df["Market_Channel_Consistency_OK"] = True
    df["Program_Duration_Consistency_OK"] = True
    df["Market_Channel_Program_Remark"] = "OK"

    def norm(x):
        if pd.isna(x):
            return ""
        return str(x).strip()

    def parse_duration_to_minutes(val):
        try:
            parts = str(val).split(":")
            if len(parts) < 2:
                return None
            h, m, s = int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) == 3 else 0
            return h * 60 + m + s / 60
        except Exception:
            return None

    reference_markets = set()
    reference_channels = set()
    if reference_df is not None:
        if "Market" in reference_df.columns:
            reference_markets.update(reference_df["Market"].dropna().astype(str).str.strip().unique())
        if "TV-Channel" in reference_df.columns:
            reference_channels.update(reference_df["TV-Channel"].dropna().astype(str).str.strip().unique())

    for idx, row in df.iterrows():
        market = norm(row.get("Market", ""))
        channel = norm(row.get("TV-Channel", ""))
        program = norm(row.get("Program Title", "")) or norm(row.get("Combined", ""))
        duration_min = parse_duration_to_minutes(row.get("Duration", ""))

        remarks = []
        ok1 = True
        ok2 = True

        if not market:
            ok1 = False
            remarks.append("Missing Market")
        elif reference_markets and market not in reference_markets:
            ok1 = False
            remarks.append(f"Unexpected Market '{market}'")

        if not channel:
            ok1 = False
            remarks.append("Missing TV-Channel")
        elif reference_channels and channel not in reference_channels:
            ok1 = False
            remarks.append(f"Unexpected TV-Channel '{channel}'")

        if not program:
            ok2 = False
            remarks.append("Missing Program Title")

        if duration_min is None:
            ok2 = False
            remarks.append("Invalid Duration")

        df.at[idx, "Market_Channel_Consistency_OK"] = ok1
        df.at[idx, "Program_Duration_Consistency_OK"] = ok2
        df.at[idx, "Market_Channel_Program_Remark"] = "; ".join(remarks) if remarks else "OK"

    return df

# -----------------------------------------------------------
# 10️⃣ Domestic Market Coverage Check
def domestic_market_coverage_check(df_worksheet, reference_df=None, debug_rows=10):
    df = df_worksheet.copy()
    df["Domestic_Market_Coverage_OK"] = True
    df["Domestic_Market_Remark"] = ""

    DOMESTIC_MAP = {
        "bundesliga": ["germany", "deutschland"],
        "premier league": ["united kingdom", "england"],
        "la liga": ["spain"],
        "serie a": ["italy"],
        "ligue 1": ["france"],
    }

    for idx, row in df.iterrows():
        comp = str(row.get("Competition", "")).lower()
        market = str(row.get("Market", "")).lower()
        progtype = str(row.get("Type of Program", "")).lower()

        domestic_markets = []
        for key, vals in DOMESTIC_MAP.items():
            if key in comp:
                domestic_markets = vals
                break
        if domestic_markets and any(k in progtype for k in ["live", "broadcast", "direct"]) and market not in domestic_markets:
            df.at[idx, "Domestic_Market_Coverage_OK"] = False
            df.at[idx, "Domestic_Market_Remark"] = f"Missing domestic live coverage for {market}"
    return df

# -----------------------------------------------------------
# 11️⃣ Rates & Ratings Check
# --------------------------------------------
def rates_and_ratings_check(df):
    """
    Rates and Ratings QC Check
    Outputs two columns:
      - Rates_Ratings_QC_OK (True/False)
      - Rates_Ratings_QC_Remark
    """
    print("\n--- Running Rates and Ratings Check ---")

    if 'Source' not in df.columns:
        df['Source'] = None
    if 'TVR% 3+' not in df.columns:
        df['TVR% 3+'] = None
    if "CPT's [Euro]" not in df.columns and "Spot price in Euro [30 sec.]" in df.columns:
        df["CPT's [Euro]"] = df["Spot price in Euro [30 sec.]"]

    df["Rates_Ratings_QC_OK"] = True
    df["Rates_Ratings_QC_Remark"] = ""

    # 1️⃣ Source overlap
    overlap_rows = []
    grouped = df.groupby(["TV-Channel", "Date"], dropna=False)
    for (channel, date), group in grouped:
        sources = group["Source"].dropna().unique().tolist()
        if "Meter" in sources and any(s not in ["Meter", None] for s in sources):
            overlap_rows.extend(group.index.tolist())

    df.loc[overlap_rows, "Rates_Ratings_QC_OK"] = False
    df.loc[overlap_rows, "Rates_Ratings_QC_Remark"] = "Meter and Non-Meter overlap"

    # 2️⃣ Linear vs OTT conflict
    if "Type of program" in df.columns:
        ott_mask = df["TV-Channel"].astype(str).str.contains("OTT", case=False, na=False)
        linear_mask = df["TV-Channel"].astype(str).str.contains("HD|TV", case=False, na=False)
        both_mask = ott_mask & linear_mask
        df.loc[both_mask, "Rates_Ratings_QC_OK"] = False
        df.loc[both_mask, "Rates_Ratings_QC_Remark"] = "Channel classified as both Linear and OTT"

    # 3️⃣ Missing rate/rating values
    invalid_rates = df[df["CPT's [Euro]"].astype(str).isin(["", "nan", "None"])]
    invalid_ratings = df[df["TVR% 3+"].astype(str).isin(["", "nan", "None"])]

    df.loc[invalid_rates.index, "Rates_Ratings_QC_OK"] = False
    df.loc[invalid_rates.index, "Rates_Ratings_QC_Remark"] = "Missing rate values"

    df.loc[invalid_ratings.index, "Rates_Ratings_QC_OK"] = False
    df.loc[invalid_ratings.index, "Rates_Ratings_QC_Remark"] = "Missing audience ratings"

    total = len(df)
    failed = (~df["Rates_Ratings_QC_OK"]).sum()
    print(f"Rates & Ratings QC Summary: {failed}/{total} failed ({(failed/total)*100:.2f}%)")

    return df

# -----------------------------------------------------------
# 12️⃣ Comparison of Duplicated Markets
def duplicated_markets_check(df):
    """
    Comparison of Duplicated Markets Check
    Outputs two columns:
      - Duplicated_Market_Check_OK (True/False)
      - Duplicated_Market_Check (remark)
    """
    print("\n--- Running Comparison of Duplicated Markets Check ---")

    for col in ["Market", "TV-Channel", "Duration"]:
        if col not in df.columns:
            df["Duplicated_Market_Check_OK"] = False
            df["Duplicated_Market_Check"] = f"Missing required column: {col}"
            print(f"⚠️ Missing required column: {col}. Skipping duplicated markets check.")
            return df

    def duration_to_hours(d):
        try:
            if pd.isna(d):
                return 0
            parts = str(d).split(":")
            h, m, s = (int(parts[i]) if i < len(parts) else 0 for i in range(3))
            return h + m/60 + s/3600
        except:
            return 0

    df["Duration_Hours"] = df["Duration"].apply(duration_to_hours)
    df["Duplicated_Market_Check_OK"] = True
    df["Duplicated_Market_Check"] = "Not Applicable"

    dup_channels = df.groupby("TV-Channel")["Market"].nunique()
    dup_channels = dup_channels[dup_channels > 1].index

    count_diff_threshold = 0.2
    duration_diff_threshold = 0.2

    for ch in dup_channels:
        subset = df[df["TV-Channel"] == ch]
        stats = subset.groupby("Market").agg(
            entry_count=("TV-Channel", "count"),
            total_duration=("Duration_Hours", "sum")
        ).reset_index()

        max_count, min_count = stats["entry_count"].max(), stats["entry_count"].min()
        max_dur, min_dur = stats["total_duration"].max(), stats["total_duration"].min()
        count_diff = abs(max_count - min_count) / max_count if max_count else 0
        dur_diff = abs(max_dur - min_dur) / max_dur if max_dur else 0

        if count_diff > count_diff_threshold or dur_diff > duration_diff_threshold:
            remark = f"Inconsistent across markets (count diff={count_diff:.0%}, duration diff={dur_diff:.0%})"
            df.loc[df["TV-Channel"] == ch, "Duplicated_Market_Check_OK"] = False
        else:
            remark = "Consistent across markets"

        df.loc[df["TV-Channel"] == ch, "Duplicated_Market_Check"] = remark

    total_checked = len(dup_channels)
    failed = (~df["Duplicated_Market_Check_OK"]).sum()
    print(f"Duplicated Markets checked: {total_checked}, Failed: {failed}")

    return df
# -----------------------------------------------------------
# 13️⃣ Country & Channel IDs Check
def country_channel_id_check(df):
    """
    Ensures that each channel and market is mapped to a single, consistent ID.
    Outputs two columns:
      - Market_Channel_ID_OK (True/False)
      - Market_Channel_ID_Remark (string)
    """

    df_result = df.copy()
    df_result["Market_Channel_ID_OK"] = True
    df_result["Market_Channel_ID_Remark"] = ""

    def norm(x):
        return str(x).strip() if pd.notna(x) else ""

    # Maps to track consistency
    channel_id_map = {}
    market_id_map = {}

    for idx, row in df_result.iterrows():
        channel = norm(row.get("TV-Channel"))
        channel_id = norm(row.get("Channel ID"))
        market = norm(row.get("Market"))
        market_id = norm(row.get("Market ID"))

        remarks = []
        ok = True

        # ✅ Check 1 – Same channel shouldn't have multiple Channel IDs
        if channel:
            if channel in channel_id_map and channel_id_map[channel] != channel_id:
                remarks.append(
                    f"Channel '{channel}' has multiple IDs ({channel_id_map[channel]} vs {channel_id})"
                )
                ok = False
            else:
                channel_id_map[channel] = channel_id

        # ✅ Check 2 – Same market shouldn't have multiple Market IDs
        if market:
            if market in market_id_map and market_id_map[market] != market_id:
                remarks.append(
                    f"Market '{market}' has multiple IDs ({market_id_map[market]} vs {market_id})"
                )
                ok = False
            else:
                market_id_map[market] = market_id

        # ✅ Check 3 – Same Channel ID shouldn't be used for multiple channels
        if channel_id and list(channel_id_map.values()).count(channel_id) > 1:
            remarks.append(f"Channel ID '{channel_id}' assigned to multiple channels")
            ok = False

        # ✅ Check 4 – Same Market ID shouldn't be used for multiple markets
        if market_id and list(market_id_map.values()).count(market_id) > 1:
            remarks.append(f"Market ID '{market_id}' assigned to multiple markets")
            ok = False

        # ✅ Write results
        df_result.at[idx, "Market_Channel_ID_OK"] = ok
        df_result.at[idx, "Market_Channel_ID_Remark"] = "; ".join(remarks) if remarks else "OK"

    return df_result

# -----------------------------------------------------------
# 14️⃣ Client Data / LSTV / OTT Check (corrected)
def client_lstv_ott_check(df_worksheet, project_config=None):
    """
    Checks:
      - Market and Channel ID consistency
      - Inclusion of Client Data, LSTV, OTT sources
    Returns:
      df with:
        - Client_LSTV_OTT_OK (True/False)
        - Client_LSTV_OTT_Remark
    """

    df = df_worksheet.copy()
    df["Client_LSTV_OTT_OK"] = True
    df["Client_LSTV_OTT_Remark"] = ""

    # --- 1️⃣ Market / Channel ID consistency ---
    if "Market ID" in df.columns and "Channel ID" in df.columns:
        # Identify Channel IDs belonging to multiple Market IDs
        multi_market = df.groupby("Channel ID")["Market ID"].nunique()
        multi_market_channels = multi_market[multi_market > 1].index.tolist()

        # Identify Market IDs belonging to multiple Channel IDs
        multi_channel = df.groupby("Market ID")["Channel ID"].nunique()
        multi_channel_ids = multi_channel[multi_channel > 1].index.tolist()
    else:
        multi_market_channels = []
        multi_channel_ids = []

    # --- 2️⃣ Client / LSTV / OTT inclusion ---
    pay_free_col = "Pay/Free TV" if "Pay/Free TV" in df.columns else None

    # Define expected sources
    expected_sources = ["lstv", "client", "ott"]

    for idx, row in df.iterrows():
        remarks = []
        ok = True

        # Market / Channel mapping issues
        if row.get("Channel ID") in multi_market_channels:
            ok = False
            remarks.append("Channel assigned to multiple Market IDs")

        if row.get("Market ID") in multi_channel_ids:
            ok = False
            remarks.append("Market ID assigned to multiple Channel IDs")

        # Client / LSTV / OTT source checks
        if pay_free_col:
            val = str(row.get(pay_free_col, "")).strip().lower()
            # Only mark False if none of the expected sources are present
            if not any(source in val for source in expected_sources):
                ok = False
                remarks.append(f"Missing required source (Client/LSTV/OTT): {row.get(pay_free_col, '')}")

        # Write results
        df.at[idx, "Client_LSTV_OTT_OK"] = ok
        df.at[idx, "Client_LSTV_OTT_Remark"] = "; ".join(remarks) if remarks else "OK"

    return df
# -----------------------------------------------------------
# ✅ Excel Coloring for True/False checks
def color_excel(output_path, df):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    wb = load_workbook(output_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col_map = {name: idx+1 for idx, name in enumerate(headers)}

    qc_columns = [col for col in df.columns if col.endswith("_OK")]

    for col_name in qc_columns:
        if col_name in col_map:
            col_idx = col_map[col_name]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                val = cell.value
                if val in [True, "True"]:
                    cell.fill = GREEN_FILL
                elif val in [False, "False"]:
                    cell.fill = RED_FILL

    wb.save(output_path)
# -----------------------------------------------------------
# Summary Sheet
def generate_summary_sheet(output_path, df):
    wb = load_workbook(output_path)
    if "Summary" in wb.sheetnames: del wb["Summary"]
    ws = wb.create_sheet("Summary")

    qc_columns = [col for col in df.columns if "_OK" in col]
    summary_data = []
    for col in qc_columns:
        total = len(df)
        passed = df[col].sum() if df[col].dtype==bool else sum(df[col]=="True")
        summary_data.append([col, total, passed, total - passed])

    summary_df = pd.DataFrame(summary_data, columns=["Check", "Total", "Passed", "Failed"])
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(r)
    wb.save(output_path)